import io
import json
from datetime import datetime
from fastapi import APIRouter, Depends, HTTPException, Request
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from database import get_db
from models import (
    User, Assessment, AssessmentAnswer, CategoryScore,
    Question, QuestionOption, QuestionCategory
)
from middleware.auth_middleware import get_current_user
from utils.scoring import get_risk_label_vi

router = APIRouter(prefix="/api/export", tags=["Export"])


@router.get("/excel/{assessment_id}")
def export_excel(
    assessment_id: str,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Export assessment results to Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    assessment = db.query(Assessment).filter(
        Assessment.id == assessment_id,
        Assessment.user_id == current_user.id,
        Assessment.status == "completed"
    ).first()
    if not assessment:
        raise HTTPException(status_code=404, detail="Không tìm thấy đánh giá")
    
    wb = Workbook()
    
    # Sheet 1: Overview
    ws = wb.active
    ws.title = "Tổng quan"
    
    header_font = Font(bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    sub_header_font = Font(bold=True, size=11)
    
    ws.merge_cells("A1:D1")
    ws["A1"] = "BÁO CÁO ĐÁNH GIÁ NGUY CƠ CHÁY NỔ"
    ws["A1"].font = header_font
    ws["A1"].fill = header_fill
    ws["A1"].alignment = Alignment(horizontal="center")
    
    info = [
        ("Tên cơ sở:", assessment.facility_name),
        ("Loại hình:", assessment.facility_type or "N/A"),
        ("Địa chỉ:", assessment.facility_address or "N/A"),
        ("Diện tích:", f"{assessment.facility_area or 'N/A'} m²"),
        ("Số người:", str(assessment.worker_count or "N/A")),
        ("Ngày đánh giá:", assessment.completed_at.strftime("%d/%m/%Y") if assessment.completed_at else "N/A"),
        ("Tổng điểm:", f"{assessment.total_score}/{assessment.max_possible_score}"),
        ("Tỷ lệ an toàn:", f"{assessment.risk_percentage}%"),
        ("Mức nguy cơ:", get_risk_label_vi(assessment.risk_level)),
    ]
    
    for i, (label, value) in enumerate(info, start=3):
        ws[f"A{i}"] = label
        ws[f"A{i}"].font = sub_header_font
        ws[f"B{i}"] = value
    
    # Sheet 2: Category Scores
    ws2 = wb.create_sheet("Điểm theo nhóm")
    headers = ["Nhóm đánh giá", "Điểm đạt", "Điểm tối đa", "Tỷ lệ %", "Mức nguy cơ"]
    for j, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=j, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
    
    cat_scores = db.query(CategoryScore).filter(
        CategoryScore.assessment_id == assessment_id
    ).all()
    
    for i, cs in enumerate(cat_scores, start=2):
        cat = db.query(QuestionCategory).get(cs.category_id)
        ws2.cell(row=i, column=1, value=cat.name if cat else "")
        ws2.cell(row=i, column=2, value=cs.score_obtained)
        ws2.cell(row=i, column=3, value=cs.max_score)
        ws2.cell(row=i, column=4, value=cs.percentage)
        ws2.cell(row=i, column=5, value=get_risk_label_vi(cs.risk_level))
    
    # Sheet 3: Detailed answers
    ws3 = wb.create_sheet("Chi tiết câu trả lời")
    headers3 = ["Nhóm", "Câu hỏi", "Trả lời", "Điểm"]
    for j, h in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=j, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
    
    answers = db.query(AssessmentAnswer).filter(
        AssessmentAnswer.assessment_id == assessment_id
    ).all()
    
    for i, ans in enumerate(answers, start=2):
        q = db.query(Question).get(ans.question_id)
        opt = db.query(QuestionOption).get(ans.selected_option_id) if ans.selected_option_id else None
        cat = db.query(QuestionCategory).get(q.category_id) if q else None
        
        ws3.cell(row=i, column=1, value=cat.name if cat else "")
        ws3.cell(row=i, column=2, value=q.question_text if q else "")
        ws3.cell(row=i, column=3, value=opt.option_text if opt else "")
        ws3.cell(row=i, column=4, value=ans.score_obtained)
    
    # Auto-width columns
    for sheet in [ws, ws2, ws3]:
        for col in sheet.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            sheet.column_dimensions[col[0].column_letter].width = min(max_length + 3, 50)
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"FRAS_Report_{assessment.facility_name}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/data/{assessment_id}")
def export_data(
    assessment_id: str,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Export assessment data as JSON for client-side PDF generation."""
    assessment = db.query(Assessment).filter(
        Assessment.id == assessment_id,
        Assessment.user_id == current_user.id,
        Assessment.status == "completed"
    ).first()
    if not assessment:
        raise HTTPException(status_code=404, detail="Không tìm thấy đánh giá")
    
    cat_scores = db.query(CategoryScore).filter(
        CategoryScore.assessment_id == assessment_id
    ).all()
    
    cat_data = []
    for cs in cat_scores:
        cat = db.query(QuestionCategory).get(cs.category_id)
        cat_data.append({
            "category_name": cat.name if cat else "",
            "icon": cat.icon if cat else "",
            "score_obtained": cs.score_obtained,
            "max_score": cs.max_score,
            "percentage": cs.percentage,
            "risk_level": cs.risk_level,
        })
    
    answers = db.query(AssessmentAnswer).filter(
        AssessmentAnswer.assessment_id == assessment_id
    ).all()
    
    answer_data = []
    for ans in answers:
        q = db.query(Question).get(ans.question_id)
        opt = db.query(QuestionOption).get(ans.selected_option_id) if ans.selected_option_id else None
        cat = db.query(QuestionCategory).get(q.category_id) if q else None
        answer_data.append({
            "category": cat.name if cat else "",
            "question": q.question_text if q else "",
            "answer": opt.option_text if opt else "",
            "score": ans.score_obtained,
        })
    
    ai_analysis = None
    if assessment.ai_analysis:
        try:
            ai_analysis = json.loads(assessment.ai_analysis)
        except json.JSONDecodeError:
            ai_analysis = {"raw": assessment.ai_analysis}
    
    return {
        "facility_name": assessment.facility_name,
        "facility_type": assessment.facility_type,
        "facility_address": assessment.facility_address,
        "facility_area": assessment.facility_area,
        "worker_count": assessment.worker_count,
        "total_score": assessment.total_score,
        "max_possible_score": assessment.max_possible_score,
        "risk_percentage": assessment.risk_percentage,
        "risk_level": assessment.risk_level,
        "completed_at": assessment.completed_at.isoformat() if assessment.completed_at else None,
        "category_scores": cat_data,
        "answers": answer_data,
        "ai_analysis": ai_analysis,
        "user_name": assessment.user.full_name if assessment.user else "",
        "organization": assessment.user.organization if assessment.user else "",
    }


@router.get("/qrcode/{assessment_id}")
def generate_qrcode(
    assessment_id: str,
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Generate a QR code for an assessment result."""
    assessment = db.query(Assessment).filter(
        Assessment.id == assessment_id,
        Assessment.user_id == current_user.id,
        Assessment.status == "completed",
    ).first()
    if not assessment:
        raise HTTPException(status_code=404, detail="Không tìm thấy đánh giá")

    import qrcode
    from qrcode.constants import ERROR_CORRECT_H

    base_url = str(request.base_url).rstrip("/")
    url = f"{base_url}/result.html?id={assessment_id}"

    qr = qrcode.QRCode(version=1, error_correction=ERROR_CORRECT_H, box_size=10, border=4)
    qr.add_data(url)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")

    output = io.BytesIO()
    img.save(output, format="PNG")
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="image/png",
        headers={"Content-Disposition": f"inline; filename=qrcode_{assessment_id[:8]}.png"},
    )

