"""
Excel Import Router
Handles bulk facility/user creation from Excel files and template download.
Optimized for large imports (up to 50,000+ accounts).
"""
import io
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from database import get_db
from models import User
from middleware.auth_middleware import get_current_user, hash_password
from middleware.rbac import require_role
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

router = APIRouter(prefix="/api/admin/users", tags=["User Import"])

DEFAULT_PASSWORD = "Pc07@123"
BATCH_SIZE = 500  # Commit every 500 rows for performance


@router.get("/import-template")
def download_import_template(
    current_user: User = Depends(get_current_user),
):
    """Download Excel template for facility/user import."""
    require_role("admin", "superadmin")(current_user)

    wb = Workbook()
    ws = wb.active
    ws.title = "Danh sách cơ sở"

    # Headers
    headers = [
        "Mã số cơ sở", "Tên cơ sở", "Email", "Số điện thoại",
        "Tỉnh", "Phường/Xã", "Vĩ độ", "Kinh độ", "Loại hình", "Mật khẩu"
    ]
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Description row
    descriptions = [
        "Mã đăng nhập (13 ký tự)",
        "VD: Cơ sở kinh doanh ABC",
        "VD: coso@email.com",
        "VD: 0901234567",
        "VD: Bắc Ninh",
        "VD: Phường Vũ Ninh",
        "VD: 21.186",
        "VD: 106.076",
        "VD: industrial,warehouse",
        f"Mặc định: {DEFAULT_PASSWORD}",
    ]
    desc_font = Font(italic=True, color="808080")
    for col, desc in enumerate(descriptions, 1):
        cell = ws.cell(row=2, column=col, value=desc)
        cell.font = desc_font

    # Example row
    example = [
        "1234567890123",
        "Cơ sở sản xuất XYZ",
        "xyz@congty.vn",
        "0901234567",
        "Bắc Ninh",
        "Phường Vũ Ninh",
        "21.186",
        "106.076",
        "industrial,warehouse",
        DEFAULT_PASSWORD,
    ]
    for col, val in enumerate(example, 1):
        ws.cell(row=3, column=col, value=val)

    # Column widths
    widths = [18, 28, 25, 16, 18, 20, 12, 12, 28, 16]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + col)].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=fras_import_facility_template.xlsx"
        },
    )


@router.post("/import-excel")
async def import_users_from_excel(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Import facilities/users from Excel file. Optimized for large files (50K+)."""
    require_role("admin", "superadmin")(current_user)

    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(
            status_code=400,
            detail="Chỉ hỗ trợ file Excel (.xlsx, .xls)",
        )

    try:
        contents = await file.read()
        wb = load_workbook(io.BytesIO(contents), read_only=True)
        ws = wb.active
    except Exception:
        raise HTTPException(
            status_code=400,
            detail="Không thể đọc file Excel. Vui lòng kiểm tra định dạng file.",
        )

    # Pre-load existing facility codes and emails to avoid N+1 queries
    existing_codes = set(
        row[0] for row in db.query(User.facility_code).filter(
            User.facility_code.isnot(None)
        ).all()
    )
    existing_emails = set(
        row[0] for row in db.query(User.email).all()
    )

    created = 0
    skipped = 0
    errors = []
    rows_processed = 0
    batch_users = []

    # Hash the default password once for reuse
    default_hash = hash_password(DEFAULT_PASSWORD)

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not any(row):
            continue

        # Skip description/example rows (check if first cell looks like a description)
        first_cell = str(row[0]).strip() if row[0] else ""
        if first_cell.startswith("Mã đăng nhập") or first_cell == "":
            continue

        facility_code = first_cell
        full_name = str(row[1]).strip() if len(row) > 1 and row[1] else ""
        email = str(row[2]).strip() if len(row) > 2 and row[2] else ""
        phone = str(row[3]).strip() if len(row) > 3 and row[3] else ""
        province = str(row[4]).strip() if len(row) > 4 and row[4] else ""
        ward = str(row[5]).strip() if len(row) > 5 and row[5] else ""
        latitude = None
        longitude = None
        try:
            latitude = float(row[6]) if len(row) > 6 and row[6] else None
        except (ValueError, TypeError):
            pass
        try:
            longitude = float(row[7]) if len(row) > 7 and row[7] else None
        except (ValueError, TypeError):
            pass
        facility_types = str(row[8]).strip() if len(row) > 8 and row[8] else ""
        password = str(row[9]).strip() if len(row) > 9 and row[9] else ""

        rows_processed += 1

        # Validation
        if not facility_code or len(facility_code) < 3:
            if len(errors) < 100:  # Cap error messages to avoid huge responses
                errors.append(f"Dòng {row_idx}: Mã cơ sở không hợp lệ ({facility_code})")
            continue

        if not full_name or len(full_name) < 2:
            if len(errors) < 100:
                errors.append(f"Dòng {row_idx}: Tên cơ sở quá ngắn ({full_name})")
            continue

        # Check duplicate using in-memory sets (fast!)
        if facility_code in existing_codes:
            skipped += 1
            continue
        
        # Generate email if not provided
        if not email or "@" not in email:
            email = f"{facility_code}@fras.local"
        
        if email in existing_emails:
            skipped += 1
            continue

        # Use pre-hashed default password or hash custom password
        if not password or len(password) < 6 or password == DEFAULT_PASSWORD:
            pw_hash = default_hash
        else:
            pw_hash = hash_password(password)

        try:
            user = User(
                email=email,
                facility_code=facility_code,
                password_hash=pw_hash,
                full_name=full_name,
                organization=full_name,
                phone=phone if phone else None,
                province=province if province else None,
                ward=ward if ward else None,
                latitude=latitude,
                longitude=longitude,
                facility_types=facility_types if facility_types else None,
                role="user",
            )
            batch_users.append(user)
            existing_codes.add(facility_code)
            existing_emails.add(email)
            created += 1

            # Batch commit every BATCH_SIZE rows
            if len(batch_users) >= BATCH_SIZE:
                db.bulk_save_objects(batch_users)
                db.commit()
                batch_users = []

        except Exception as e:
            if len(errors) < 100:
                errors.append(f"Dòng {row_idx}: Lỗi tạo tài khoản - {str(e)}")

    # Final batch commit
    if batch_users:
        db.bulk_save_objects(batch_users)
        db.commit()

    wb.close()

    return {
        "status": "ok",
        "created": created,
        "skipped": skipped,
        "errors": errors[:100],  # Cap at 100 error messages
        "total_errors": len(errors),
        "total_processed": rows_processed,
        "message": f"Đã tạo {created} tài khoản cơ sở mới"
        + (f", bỏ qua {skipped} (trùng mã)" if skipped else "")
        + (f", {len(errors)} lỗi" if errors else ""),
    }
